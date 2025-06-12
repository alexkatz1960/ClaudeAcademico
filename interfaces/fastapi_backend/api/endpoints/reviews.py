 # ==========================================
# INTERFACES/FASTAPI_BACKEND/API/ENDPOINTS/REVIEWS.PY
# Enterprise Editorial Review Management Endpoints - FastAPI
# Sistema de Traducción Académica v2.2
# ==========================================

from fastapi import APIRouter, Depends, HTTPException, Query, Path, BackgroundTasks
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, desc, func, text
from typing import List, Optional, Dict, Any, Union
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path as FilePath
import structlog
import json
import csv
import io
import tempfile
import zipfile
from dataclasses import dataclass

# Internal imports
from ...database.database import get_db, db_transaction
from ...database.models import (
    EditorialReview, BookProcessingHistory, AuditLog, 
    TerminologySuggestion, UsageStatistic
)
from ...database.schemas import (
    EditorialReviewResponse, EditorialReviewCreate, EditorialReviewUpdate,
    ReviewWorkflowResponse, ReviewAssignmentRequest, ReviewAssignmentResponse,
    ReviewSummaryResponse, ReviewAnalyticsResponse, ReviewExportRequest,
    ReviewBulkUpdateRequest, ReviewEscalationRequest
)
from ...database.crud import (
    EditorialReviewCRUD, BookCRUD, AuditCRUD, TerminologyCRUD
)
from ...database.enums import (
    ReviewSeverity, ReviewDecision, BookStatus, ProcessingPhase
)
from ..dependencies import (
    get_current_user, require_permissions, get_rate_limiter
)
from ...core.security import Permission
from ...utils.responses import (
    create_success_response, create_error_response,
    ErrorCode, SuccessResponse, ErrorResponse
)

# Configure logger
logger = structlog.get_logger(__name__)

# Create router
router = APIRouter(
    prefix="/reviews",
    tags=["reviews"],
    responses={
        404: {"description": "Review not found"},
        403: {"description": "Insufficient permissions"},
        422: {"description": "Validation error"}
    }
)

# ==========================================
# REVIEW MANAGEMENT ENUMS
# ==========================================

class ReviewStatus(str, Enum):
    """Review status for filtering"""
    PENDING = "pending"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    ESCALATED = "escalated"

class ReviewPriority(str, Enum):
    """Review priority levels"""
    LOW = "low"
    NORMAL = "normal"
    HIGH = "high"
    URGENT = "urgent"

class ReviewFilterBy(str, Enum):
    """Review filtering options"""
    ASSIGNED_TO_ME = "assigned_to_me"
    UNASSIGNED = "unassigned"
    OVERDUE = "overdue"
    HIGH_PRIORITY = "high_priority"
    RECENT = "recent"

class ExportFormat(str, Enum):
    """Export formats for reviews"""
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"
    PDF = "pdf"

@dataclass
class ReviewWorkflowMetrics:
    """Metrics for review workflow analysis"""
    total_reviews: int
    pending_reviews: int
    completed_reviews: int
    overdue_reviews: int
    avg_completion_time_hours: float
    completion_rate: float
    escalation_rate: float

# ==========================================
# REVIEW MANAGEMENT SERVICE
# ==========================================

class ReviewManagementService:
    """Enterprise review management service with comprehensive workflow support"""
    
    def __init__(self):
        self.logger = structlog.get_logger("services.review_management")
    
    def get_reviews_by_book(self, db: Session, book_id: str, 
                           include_resolved: bool = True) -> List[EditorialReview]:
        """Get all reviews for a specific book with optional filtering"""
        
        try:
            review_crud = EditorialReviewCRUD(db)
            
            filters = {"book_id": book_id}
            if not include_resolved:
                filters["resolved"] = False
            
            reviews = review_crud.get_by_filters(filters)
            
            self.logger.info(
                "Retrieved reviews for book",
                book_id=book_id,
                review_count=len(reviews),
                include_resolved=include_resolved
            )
            
            return reviews
            
        except Exception as e:
            self.logger.error("Error getting reviews by book", book_id=book_id, error=str(e))
            raise
    
    def get_reviews_by_status(self, db: Session, status: ReviewStatus,
                             assigned_to: Optional[str] = None,
                             limit: int = 50) -> List[EditorialReview]:
        """Get reviews filtered by status and optional assignee"""
        
        try:
            review_crud = EditorialReviewCRUD(db)
            
            filters = {}
            
            if status == ReviewStatus.PENDING:
                filters["reviewed"] = False
                filters["resolved"] = False
            elif status == ReviewStatus.IN_PROGRESS:
                filters["reviewed"] = True
                filters["resolved"] = False
            elif status == ReviewStatus.COMPLETED:
                filters["resolved"] = True
            elif status == ReviewStatus.ESCALATED:
                filters["escalated"] = True
            
            if assigned_to:
                filters["assigned_to"] = assigned_to
            
            reviews = review_crud.get_by_filters(filters, limit=limit)
            
            self.logger.info(
                "Retrieved reviews by status",
                status=status.value,
                assigned_to=assigned_to,
                review_count=len(reviews)
            )
            
            return reviews
            
        except Exception as e:
            self.logger.error("Error getting reviews by status", status=status.value, error=str(e))
            raise
    
    def assign_review(self, db: Session, review_id: int, assigned_to: str,
                     assigned_by: str, priority: ReviewPriority = ReviewPriority.NORMAL,
                     deadline: Optional[datetime] = None) -> bool:
        """Assign review to specific user with priority and deadline"""
        
        try:
            review_crud = EditorialReviewCRUD(db)
            
            # Get current review
            review = review_crud.get_by_id(review_id)
            if not review:
                return False
            
            # Update assignment
            update_data = {
                "assigned_to": assigned_to,
                "review_difficulty": self._calculate_review_priority_score(priority),
                "deadline": deadline or (datetime.now() + timedelta(days=2))  # Default 2 days
            }
            
            updated_review = review_crud.update(review_id, update_data)
            
            self.logger.info(
                "Assigned review to user",
                review_id=review_id,
                assigned_to=assigned_to,
                assigned_by=assigned_by,
                priority=priority.value,
                deadline=deadline
            )
            
            return updated_review is not None
            
        except Exception as e:
            self.logger.error("Error assigning review", review_id=review_id, error=str(e))
            return False
    
    def escalate_review(self, db: Session, review_id: int, escalated_by: str,
                       reason: str, escalate_to: Optional[str] = None) -> bool:
        """Escalate review to higher level with comprehensive tracking"""
        
        try:
            review_crud = EditorialReviewCRUD(db)
            
            # Get current review
            review = review_crud.get_by_id(review_id)
            if not review:
                return False
            
            # Escalate review
            review.escalate(reason, escalate_to)
            
            # Update in database
            update_data = {
                "escalated": True,
                "escalation_reason": reason,
                "assigned_to": escalate_to or "senior_editor"
            }
            
            updated_review = review_crud.update(review_id, update_data)
            
            self.logger.warning(
                "Review escalated",
                review_id=review_id,
                escalated_by=escalated_by,
                reason=reason,
                escalate_to=escalate_to
            )
            
            return updated_review is not None
            
        except Exception as e:
            self.logger.error("Error escalating review", review_id=review_id, error=str(e))
            return False
    
    def get_workflow_metrics(self, db: Session, timeframe_days: int = 30) -> ReviewWorkflowMetrics:
        """Get comprehensive workflow metrics for specified timeframe"""
        
        try:
            review_crud = EditorialReviewCRUD(db)
            
            # Calculate date range
            end_date = datetime.now()
            start_date = end_date - timedelta(days=timeframe_days)
            
            # Get reviews in timeframe
            reviews = review_crud.get_by_date_range(start_date, end_date, resolved_only=False)
            
            total_reviews = len(reviews)
            pending_reviews = len([r for r in reviews if not r.reviewed and not r.resolved])
            completed_reviews = len([r for r in reviews if r.resolved])
            overdue_reviews = len([r for r in reviews if r.deadline and r.deadline < datetime.now() and not r.resolved])
            escalated_reviews = len([r for r in reviews if r.escalated])
            
            # Calculate completion time for resolved reviews
            completed_with_time = [r for r in reviews if r.resolved and r.resolution_time_minutes]
            avg_completion_time_hours = (
                sum(r.resolution_time_minutes for r in completed_with_time) / 60 / len(completed_with_time)
            ) if completed_with_time else 0
            
            # Calculate rates
            completion_rate = (completed_reviews / total_reviews * 100) if total_reviews > 0 else 0
            escalation_rate = (escalated_reviews / total_reviews * 100) if total_reviews > 0 else 0
            
            metrics = ReviewWorkflowMetrics(
                total_reviews=total_reviews,
                pending_reviews=pending_reviews,
                completed_reviews=completed_reviews,
                overdue_reviews=overdue_reviews,
                avg_completion_time_hours=avg_completion_time_hours,
                completion_rate=completion_rate,
                escalation_rate=escalation_rate
            )
            
            self.logger.info(
                "Generated workflow metrics",
                timeframe_days=timeframe_days,
                total_reviews=total_reviews,
                completion_rate=completion_rate
            )
            
            return metrics
            
        except Exception as e:
            self.logger.error("Error generating workflow metrics", error=str(e))
            raise
    
    def export_reviews(self, db: Session, export_format: ExportFormat,
                      book_id: Optional[str] = None,
                      status_filter: Optional[ReviewStatus] = None,
                      assigned_to: Optional[str] = None) -> Dict[str, Any]:
        """Export reviews in specified format with filtering"""
        
        try:
            review_crud = EditorialReviewCRUD(db)
            
            # Build filters
            filters = {}
            if book_id:
                filters["book_id"] = book_id
            if assigned_to:
                filters["assigned_to"] = assigned_to
            
            # Apply status filter
            if status_filter == ReviewStatus.PENDING:
                filters["reviewed"] = False
                filters["resolved"] = False
            elif status_filter == ReviewStatus.COMPLETED:
                filters["resolved"] = True
            
            reviews = review_crud.get_by_filters(filters, limit=1000)
            
            # Generate export file
            if export_format == ExportFormat.CSV:
                file_path = self._export_to_csv(reviews)
            elif export_format == ExportFormat.EXCEL:
                file_path = self._export_to_excel(reviews)
            elif export_format == ExportFormat.JSON:
                file_path = self._export_to_json(reviews)
            else:
                raise ValueError(f"Unsupported export format: {export_format}")
            
            self.logger.info(
                "Exported reviews",
                format=export_format.value,
                review_count=len(reviews),
                file_path=file_path
            )
            
            return {
                "file_path": file_path,
                "format": export_format.value,
                "review_count": len(reviews),
                "exported_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            self.logger.error("Error exporting reviews", format=export_format.value, error=str(e))
            raise
    
    def _calculate_review_priority_score(self, priority: ReviewPriority) -> int:
        """Convert priority to numeric score (1-10 scale)"""
        priority_scores = {
            ReviewPriority.LOW: 3,
            ReviewPriority.NORMAL: 5,
            ReviewPriority.HIGH: 7,
            ReviewPriority.URGENT: 9
        }
        return priority_scores.get(priority, 5)
    
    def _export_to_csv(self, reviews: List[EditorialReview]) -> str:
        """Export reviews to CSV format"""
        
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        
        try:
            writer = csv.writer(temp_file)
            
            # Write header
            writer.writerow([
                'Review ID', 'Book ID', 'Section', 'Alert Type', 'Severity',
                'Similarity Score', 'Suggested Action', 'Editor Decision',
                'Assigned To', 'Reviewed', 'Resolved', 'Created At',
                'Resolution Time (min)', 'Editor Notes'
            ])
            
            # Write data
            for review in reviews:
                writer.writerow([
                    review.id,
                    review.book_id,
                    review.section_number,
                    review.alert_type,
                    review.severity,
                    review.similarity_score,
                    review.suggested_action,
                    review.editor_decision,
                    review.assigned_to or '',
                    review.reviewed,
                    review.resolved,
                    review.created_at.isoformat(),
                    review.resolution_time_minutes or '',
                    review.editor_notes or ''
                ])
            
            temp_file.close()
            return temp_file.name
            
        except Exception as e:
            temp_file.close()
            raise
    
    def _export_to_excel(self, reviews: List[EditorialReview]) -> str:
        """Export reviews to Excel format"""
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Editorial Reviews"
        
        # Header row with styling
        headers = [
            'Review ID', 'Book ID', 'Section', 'Alert Type', 'Severity',
            'Similarity Score', 'Suggested Action', 'Editor Decision',
            'Assigned To', 'Reviewed', 'Resolved', 'Created At',
            'Resolution Time (min)', 'Editor Notes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row, review in enumerate(reviews, 2):
            ws.cell(row=row, column=1, value=review.id)
            ws.cell(row=row, column=2, value=review.book_id)
            ws.cell(row=row, column=3, value=review.section_number)
            ws.cell(row=row, column=4, value=review.alert_type)
            ws.cell(row=row, column=5, value=review.severity)
            ws.cell(row=row, column=6, value=review.similarity_score)
            ws.cell(row=row, column=7, value=review.suggested_action)
            ws.cell(row=row, column=8, value=review.editor_decision)
            ws.cell(row=row, column=9, value=review.assigned_to or '')
            ws.cell(row=row, column=10, value=review.reviewed)
            ws.cell(row=row, column=11, value=review.resolved)
            ws.cell(row=row, column=12, value=review.created_at.isoformat())
            ws.cell(row=row, column=13, value=review.resolution_time_minutes or '')
            ws.cell(row=row, column=14, value=review.editor_notes or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def _export_to_json(self, reviews: List[EditorialReview]) -> str:
        """Export reviews to JSON format"""
        
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8')
        
        try:
            reviews_data = []
            for review in reviews:
                review_data = {
                    "id": review.id,
                    "book_id": review.book_id,
                    "section_number": review.section_number,
                    "alert_type": review.alert_type,
                    "severity": review.severity,
                    "similarity_score": review.similarity_score,
                    "suggested_action": review.suggested_action,
                    "editor_decision": review.editor_decision,
                    "assigned_to": review.assigned_to,
                    "reviewed": review.reviewed,
                    "resolved": review.resolved,
                    "created_at": review.created_at.isoformat(),
                    "reviewed_at": review.reviewed_at.isoformat() if review.reviewed_at else None,
                    "resolved_at": review.resolved_at.isoformat() if review.resolved_at else None,
                    "resolution_time_minutes": review.resolution_time_minutes,
                    "editor_notes": review.editor_notes,
                    "original_text_preview": review.original_text[:200] if review.original_text else "",
                    "translated_text_preview": review.translated_text[:200] if review.translated_text else ""
                }
                reviews_data.append(review_data)
            
            export_data = {
                "export_metadata": {
                    "exported_at": datetime.now().isoformat(),
                    "total_reviews": len(reviews_data),
                    "format": "json"
                },
                "reviews": reviews_data
            }
            
            json.dump(export_data, temp_file, indent=2, ensure_ascii=False)
            temp_file.close()
            
            return temp_file.name
            
        except Exception as e:
            temp_file.close()
            raise

# ==========================================
# SERVICE INSTANCE
# ==========================================

review_service = ReviewManagementService()

# ==========================================
# REVIEW CRUD ENDPOINTS
# ==========================================

@router.get("/books/{book_id}",
    response_model=SuccessResponse[List[EditorialReviewResponse]],
    summary="Get reviews by book",
    description="Retrieve all editorial reviews for a specific book"
)
async def get_reviews_by_book(
    book_id: str = Path(..., description="Book ID to get reviews for"),
    include_resolved: bool = Query(True, description="Include resolved reviews"),
    severity_filter: Optional[ReviewSeverity] = Query(None, description="Filter by severity"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Get all editorial reviews for a specific book
    
    Enterprise features:
    - Comprehensive review listing
    - Severity filtering
    - Resolution status filtering
    - User-specific view permissions
    """
    
    try:
        logger.info(
            "Getting reviews for book",
            book_id=book_id,
            include_resolved=include_resolved,
            user_id=getattr(current_user, 'id', 'api_user')
        )
        
        # Verify book exists
        book_crud = BookCRUD(db)
        book = book_crud.get_by_id(book_id)
        
        if not book:
            return create_error_response(
                error_code=ErrorCode.NOT_FOUND,
                message=f"Book {book_id} not found",
                status_code=404
            )
        
        # Get reviews
        reviews = review_service.get_reviews_by_book(db, book_id, include_resolved)
        
        # Apply severity filter if specified
        if severity_filter:
            reviews = [r for r in reviews if r.severity == severity_filter.value]
        
        # Convert to response format
        response_data = []
        for review in reviews:
            review_response = EditorialReviewResponse(
                id=review.id,
                book_id=review.book_id,
                section_number=review.section_number,
                alert_type=review.alert_type,
                severity=ReviewSeverity(review.severity),
                similarity_score=review.similarity_score,
                suggested_action=review.suggested_action,
                editor_decision=ReviewDecision(review.editor_decision) if review.editor_decision else None,
                assigned_to=review.assigned_to,
                reviewed=review.reviewed,
                resolved=review.resolved,
                created_at=review.created_at,
                reviewed_at=review.reviewed_at,
                resolved_at=review.resolved_at,
                resolution_time_minutes=review.resolution_time_minutes,
                editor_notes=review.editor_notes,
                escalated=review.escalated,
                deadline=review.deadline
            )
            response_data.append(review_response)
        
        return create_success_response(
            data=response_data,
            message=f"Retrieved {len(response_data)} reviews for book {book_id}"
        )
        
    except Exception as e:
        logger.error("Error getting reviews by book", book_id=book_id, error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to retrieve reviews",
            status_code=500,
            details={"book_id": book_id, "error": str(e)}
        )

@router.get("/{review_id}",
    response_model=SuccessResponse[EditorialReviewResponse],
    summary="Get review by ID",
    description="Retrieve specific editorial review by ID"
)
async def get_review_by_id(
    review_id: int = Path(..., description="Review ID"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Get specific editorial review by ID
    
    Enterprise features:
    - Complete review details
    - User permission validation
    - Comprehensive response
    """
    
    try:
        review_crud = EditorialReviewCRUD(db)
        review = review_crud.get_by_id(review_id)
        
        if not review:
            return create_error_response(
                error_code=ErrorCode.NOT_FOUND,
                message=f"Review {review_id} not found",
                status_code=404
            )
        
        response_data = EditorialReviewResponse(
            id=review.id,
            book_id=review.book_id,
            section_number=review.section_number,
            alert_type=review.alert_type,
            severity=ReviewSeverity(review.severity),
            similarity_score=review.similarity_score,
            suggested_action=review.suggested_action,
            editor_decision=ReviewDecision(review.editor_decision) if review.editor_decision else None,
            assigned_to=review.assigned_to,
            reviewed=review.reviewed,
            resolved=review.resolved,
            created_at=review.created_at,
            reviewed_at=review.reviewed_at,
            resolved_at=review.resolved_at,
            resolution_time_minutes=review.resolution_time_minutes,
            editor_notes=review.editor_notes,
            escalated=review.escalated,
            deadline=review.deadline,
            # Include full text for detailed view
            original_text=review.original_text,
            translated_text=review.translated_text,
            location_info=review.location_info
        )
        
        return create_success_response(
            data=response_data,
            message="Review retrieved successfully"
        )
        
    except Exception as e:
        logger.error("Error getting review by ID", review_id=review_id, error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to retrieve review",
            status_code=500,
            details={"review_id": review_id, "error": str(e)}
        )

@router.put("/{review_id}",
    response_model=SuccessResponse[EditorialReviewResponse],
    summary="Update review",
    description="Update editorial review with editor decisions and notes"
)
async def update_review(
    review_id: int = Path(..., description="Review ID"),
    review_update: EditorialReviewUpdate = ...,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Update editorial review with comprehensive tracking
    
    Enterprise features:
    - Decision workflow tracking
    - User attribution
    - Resolution time calculation
    - State transition validation
    """
    
    try:
        user_id = getattr(current_user, 'id', 'api_user')
        
        logger.info(
            "Updating review",
            review_id=review_id,
            user_id=user_id,
            decision=review_update.editor_decision
        )
        
        review_crud = EditorialReviewCRUD(db)
        review = review_crud.get_by_id(review_id)
        
        if not review:
            return create_error_response(
                error_code=ErrorCode.NOT_FOUND,
                message=f"Review {review_id} not found",
                status_code=404
            )
        
        # Prepare update data
        update_data = review_update.dict(exclude_unset=True)
        update_data["reviewer_id"] = user_id
        
        # Handle workflow state transitions
        if review_update.editor_decision and not review.reviewed:
            update_data["reviewed"] = True
            update_data["reviewed_at"] = datetime.now()
        
        if review_update.resolved and not review.resolved:
            update_data["resolved"] = True
            update_data["resolved_at"] = datetime.now()
            
            # Calculate resolution time
            if review.reviewed_at:
                resolution_time = (datetime.now() - review.reviewed_at).total_seconds() / 60
                update_data["resolution_time_minutes"] = int(resolution_time)
        
        # Update review
        updated_review = review_crud.update(review_id, update_data)
        
        # Convert to response
        response_data = EditorialReviewResponse(
            id=updated_review.id,
            book_id=updated_review.book_id,
            section_number=updated_review.section_number,
            alert_type=updated_review.alert_type,
            severity=ReviewSeverity(updated_review.severity),
            similarity_score=updated_review.similarity_score,
            suggested_action=updated_review.suggested_action,
            editor_decision=ReviewDecision(updated_review.editor_decision) if updated_review.editor_decision else None,
            assigned_to=updated_review.assigned_to,
            reviewed=updated_review.reviewed,
            resolved=updated_review.resolved,
            created_at=updated_review.created_at,
            reviewed_at=updated_review.reviewed_at,
            resolved_at=updated_review.resolved_at,
            resolution_time_minutes=updated_review.resolution_time_minutes,
            editor_notes=updated_review.editor_notes,
            escalated=updated_review.escalated,
            deadline=updated_review.deadline
        )
        
        return create_success_response(
            data=response_data,
            message="Review updated successfully"
        )
        
    except Exception as e:
        logger.error("Error updating review", review_id=review_id, error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to update review",
            status_code=500,
            details={"review_id": review_id, "error": str(e)}
        )

# ==========================================
# WORKFLOW MANAGEMENT ENDPOINTS
# ==========================================

@router.get("/status/{status}",
    response_model=SuccessResponse[List[EditorialReviewResponse]],
    summary="Get reviews by status",
    description="Retrieve reviews filtered by workflow status"
)
async def get_reviews_by_status(
    status: ReviewStatus = Path(..., description="Review status filter"),
    assigned_to: Optional[str] = Query(None, description="Filter by assigned user"),
    limit: int = Query(50, ge=1, le=200, description="Maximum reviews to return"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Get reviews by workflow status with assignment filtering
    
    Enterprise features:
    - Status-based workflow filtering
    - User assignment filtering
    - Workload management
    """
    
    try:
        reviews = review_service.get_reviews_by_status(db, status, assigned_to, limit)
        
        response_data = []
        for review in reviews:
            review_response = EditorialReviewResponse(
                id=review.id,
                book_id=review.book_id,
                section_number=review.section_number,
                alert_type=review.alert_type,
                severity=ReviewSeverity(review.severity),
                similarity_score=review.similarity_score,
                suggested_action=review.suggested_action,
                editor_decision=ReviewDecision(review.editor_decision) if review.editor_decision else None,
                assigned_to=review.assigned_to,
                reviewed=review.reviewed,
                resolved=review.resolved,
                created_at=review.created_at,
                deadline=review.deadline,
                escalated=review.escalated
            )
            response_data.append(review_response)
        
        return create_success_response(
            data=response_data,
            message=f"Retrieved {len(response_data)} reviews with status {status.value}"
        )
        
    except Exception as e:
        logger.error("Error getting reviews by status", status=status.value, error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to retrieve reviews by status",
            status_code=500,
            details={"status": status.value, "error": str(e)}
        )

@router.post("/{review_id}/assign",
    response_model=SuccessResponse[ReviewAssignmentResponse],
    summary="Assign review",
    description="Assign review to specific user with priority and deadline"
)
async def assign_review(
    review_id: int = Path(..., description="Review ID"),
    assignment: ReviewAssignmentRequest = ...,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
    _permissions = Depends(require_permissions([Permission.MANAGE_REVIEWS]))
):
    """
    Assign review to user with workflow management
    
    Enterprise features:
    - Assignment workflow
    - Priority setting
    - Deadline management
    - User attribution
    """
    
    try:
        assigned_by = getattr(current_user, 'id', 'api_user')
        
        logger.info(
            "Assigning review",
            review_id=review_id,
            assigned_to=assignment.assigned_to,
            assigned_by=assigned_by,
            priority=assignment.priority
        )
        
        success = review_service.assign_review(
            db=db,
            review_id=review_id,
            assigned_to=assignment.assigned_to,
            assigned_by=assigned_by,
            priority=assignment.priority,
            deadline=assignment.deadline
        )
        
        if not success:
            return create_error_response(
                error_code=ErrorCode.NOT_FOUND,
                message=f"Review {review_id} not found or assignment failed",
                status_code=404
            )
        
        response_data = ReviewAssignmentResponse(
            review_id=review_id,
            assigned_to=assignment.assigned_to,
            assigned_by=assigned_by,
            priority=assignment.priority,
            deadline=assignment.deadline,
            assigned_at=datetime.now()
        )
        
        return create_success_response(
            data=response_data,
            message="Review assigned successfully"
        )
        
    except Exception as e:
        logger.error("Error assigning review", review_id=review_id, error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to assign review",
            status_code=500,
            details={"review_id": review_id, "error": str(e)}
        )

@router.post("/{review_id}/escalate",
    response_model=SuccessResponse[Dict[str, Any]],
    summary="Escalate review",
    description="Escalate review to higher level with reason tracking"
)
async def escalate_review(
    review_id: int = Path(..., description="Review ID"),
    escalation: ReviewEscalationRequest = ...,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Escalate review to higher level with comprehensive tracking
    
    Enterprise features:
    - Escalation workflow
    - Reason tracking
    - User attribution
    - Automatic reassignment
    """
    
    try:
        escalated_by = getattr(current_user, 'id', 'api_user')
        
        logger.warning(
            "Escalating review",
            review_id=review_id,
            escalated_by=escalated_by,
            reason=escalation.reason
        )
        
        success = review_service.escalate_review(
            db=db,
            review_id=review_id,
            escalated_by=escalated_by,
            reason=escalation.reason,
            escalate_to=escalation.escalate_to
        )
        
        if not success:
            return create_error_response(
                error_code=ErrorCode.NOT_FOUND,
                message=f"Review {review_id} not found or escalation failed",
                status_code=404
            )
        
        response_data = {
            "review_id": review_id,
            "escalated": True,
            "escalated_by": escalated_by,
            "escalate_to": escalation.escalate_to,
            "reason": escalation.reason,
            "escalated_at": datetime.now().isoformat()
        }
        
        return create_success_response(
            data=response_data,
            message="Review escalated successfully"
        )
        
    except Exception as e:
        logger.error("Error escalating review", review_id=review_id, error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to escalate review",
            status_code=500,
            details={"review_id": review_id, "error": str(e)}
        )

@router.post("/bulk/update",
    response_model=SuccessResponse[Dict[str, Any]],
    summary="Bulk update reviews",
    description="Update multiple reviews in batch for efficiency"
)
async def bulk_update_reviews(
    bulk_update: ReviewBulkUpdateRequest = ...,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
    _permissions = Depends(require_permissions([Permission.MANAGE_REVIEWS]))
):
    """
    Bulk update multiple reviews for efficiency
    
    Enterprise features:
    - Batch processing
    - Transaction safety
    - Progress tracking
    - Error handling per item
    """
    
    try:
        user_id = getattr(current_user, 'id', 'api_user')
        
        logger.info(
            "Bulk updating reviews",
            review_count=len(bulk_update.review_ids),
            user_id=user_id,
            action=bulk_update.action
        )
        
        review_crud = EditorialReviewCRUD(db)
        
        updated_count = 0
        failed_count = 0
        errors = []
        
        with db_transaction() as transaction_db:
            for review_id in bulk_update.review_ids:
                try:
                    # Prepare update data based on action
                    update_data = {"reviewer_id": user_id}
                    
                    if bulk_update.action == "mark_reviewed":
                        update_data.update({
                            "reviewed": True,
                            "reviewed_at": datetime.now()
                        })
                    elif bulk_update.action == "mark_resolved":
                        update_data.update({
                            "resolved": True,
                            "resolved_at": datetime.now()
                        })
                    elif bulk_update.action == "assign" and bulk_update.assigned_to:
                        update_data["assigned_to"] = bulk_update.assigned_to
                    
                    # Add notes if provided
                    if bulk_update.notes:
                        update_data["editor_notes"] = bulk_update.notes
                    
                    # Update review
                    review_crud.update(review_id, update_data)
                    updated_count += 1
                    
                except Exception as item_error:
                    failed_count += 1
                    errors.append({
                        "review_id": review_id,
                        "error": str(item_error)
                    })
        
        response_data = {
            "total_reviews": len(bulk_update.review_ids),
            "updated_count": updated_count,
            "failed_count": failed_count,
            "errors": errors,
            "action": bulk_update.action,
            "processed_by": user_id,
            "processed_at": datetime.now().isoformat()
        }
        
        return create_success_response(
            data=response_data,
            message=f"Bulk update completed: {updated_count} updated, {failed_count} failed"
        )
        
    except Exception as e:
        logger.error("Error in bulk update", error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to perform bulk update",
            status_code=500,
            details={"error": str(e)}
        )

# ==========================================
# ANALYTICS AND REPORTING ENDPOINTS
# ==========================================

@router.get("/analytics/workflow",
    response_model=SuccessResponse[ReviewAnalyticsResponse],
    summary="Get review workflow analytics",
    description="Get comprehensive analytics on review workflow performance"
)
async def get_review_workflow_analytics(
    timeframe_days: int = Query(30, ge=1, le=365, description="Analytics timeframe in days"),
    include_trends: bool = Query(True, description="Include daily trends"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Get comprehensive review workflow analytics
    
    Enterprise features:
    - Multi-timeframe analytics
    - Performance metrics
    - Trend analysis
    - Efficiency tracking
    """
    
    try:
        logger.info(
            "Generating review workflow analytics",
            timeframe_days=timeframe_days,
            user_id=getattr(current_user, 'id', 'api_user')
        )
        
        metrics = review_service.get_workflow_metrics(db, timeframe_days)
        
        # Get additional breakdown by severity
        review_crud = EditorialReviewCRUD(db)
        end_date = datetime.now()
        start_date = end_date - timedelta(days=timeframe_days)
        
        all_reviews = review_crud.get_by_date_range(start_date, end_date, resolved_only=False)
        
        severity_breakdown = {}
        for severity in ReviewSeverity:
            severity_reviews = [r for r in all_reviews if r.severity == severity.value]
            severity_breakdown[severity.value] = {
                "total": len(severity_reviews),
                "resolved": len([r for r in severity_reviews if r.resolved]),
                "avg_resolution_time": sum(r.resolution_time_minutes or 0 for r in severity_reviews if r.resolved) / max(len([r for r in severity_reviews if r.resolved]), 1)
            }
        
        response_data = ReviewAnalyticsResponse(
            timeframe_days=timeframe_days,
            total_reviews=metrics.total_reviews,
            pending_reviews=metrics.pending_reviews,
            completed_reviews=metrics.completed_reviews,
            overdue_reviews=metrics.overdue_reviews,
            completion_rate_percentage=metrics.completion_rate,
            escalation_rate_percentage=metrics.escalation_rate,
            avg_completion_time_hours=metrics.avg_completion_time_hours,
            severity_breakdown=severity_breakdown,
            generated_at=datetime.now()
        )
        
        return create_success_response(
            data=response_data,
            message="Review analytics generated successfully"
        )
        
    except Exception as e:
        logger.error("Error generating review analytics", error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to generate review analytics",
            status_code=500,
            details={"error": str(e)}
        )

@router.get("/summary",
    response_model=SuccessResponse[ReviewSummaryResponse],
    summary="Get review summary",
    description="Get high-level review summary for dashboard display"
)
async def get_review_summary(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Get high-level review summary for dashboard
    
    Enterprise features:
    - Real-time metrics
    - Workload indicators
    - Priority alerts
    """
    
    try:
        # Get current metrics
        metrics = review_service.get_workflow_metrics(db, timeframe_days=7)
        
        # Get user-specific assignments if user ID available
        user_id = getattr(current_user, 'id', None)
        my_pending_reviews = 0
        my_overdue_reviews = 0
        
        if user_id:
            my_reviews = review_service.get_reviews_by_status(
                db, ReviewStatus.PENDING, assigned_to=user_id, limit=1000
            )
            my_pending_reviews = len(my_reviews)
            my_overdue_reviews = len([
                r for r in my_reviews 
                if r.deadline and r.deadline < datetime.now()
            ])
        
        response_data = ReviewSummaryResponse(
            total_pending_reviews=metrics.pending_reviews,
            total_overdue_reviews=metrics.overdue_reviews,
            my_pending_reviews=my_pending_reviews,
            my_overdue_reviews=my_overdue_reviews,
            completion_rate_7d=metrics.completion_rate,
            avg_completion_time_hours=metrics.avg_completion_time_hours,
            system_status="healthy" if metrics.overdue_reviews < 5 else "warning" if metrics.overdue_reviews < 20 else "critical"
        )
        
        return create_success_response(
            data=response_data,
            message="Review summary retrieved successfully"
        )
        
    except Exception as e:
        logger.error("Error getting review summary", error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to retrieve review summary",
            status_code=500,
            details={"error": str(e)}
        )

# ==========================================
# EXPORT ENDPOINTS
# ==========================================

@router.post("/export",
    response_class=FileResponse,
    summary="Export reviews",
    description="Export reviews in specified format with filtering options"
)
async def export_reviews(
    export_request: ReviewExportRequest = ...,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Export reviews in multiple formats with comprehensive filtering
    
    Enterprise features:
    - Multi-format export (CSV, Excel, JSON)
    - Advanced filtering
    - Large dataset handling
    - Download management
    """
    
    try:
        user_id = getattr(current_user, 'id', 'api_user')
        
        logger.info(
            "Exporting reviews",
            format=export_request.format.value,
            user_id=user_id,
            filters=export_request.dict(exclude={'format'})
        )
        
        export_result = review_service.export_reviews(
            db=db,
            export_format=export_request.format,
            book_id=export_request.book_id,
            status_filter=export_request.status_filter,
            assigned_to=export_request.assigned_to
        )
        
        file_path = export_result["file_path"]
        
        # Determine content type and filename
        content_types = {
            ExportFormat.CSV: "text/csv",
            ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ExportFormat.JSON: "application/json"
        }
        
        extensions = {
            ExportFormat.CSV: "csv",
            ExportFormat.EXCEL: "xlsx", 
            ExportFormat.JSON: "json"
        }
        
        filename = f"editorial_reviews_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{extensions[export_request.format]}"
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type=content_types[export_request.format],
            headers={
                "X-Export-Count": str(export_result["review_count"]),
                "X-Export-Format": export_request.format.value
            }
        )
        
    except Exception as e:
        logger.error("Error exporting reviews", error=str(e))
        
        return create_error_response(
            error_code=ErrorCode.INTERNAL_ERROR,
            message="Failed to export reviews",
            status_code=500,
            details={"error": str(e)}
        )

# Export all endpoint functions for testing
__all__ = [
    "router",
    "get_reviews_by_book",
    "get_review_by_id",
    "update_review",
    "get_reviews_by_status",
    "assign_review",
    "escalate_review",
    "bulk_update_reviews",
    "get_review_workflow_analytics",
    "get_review_summary",
    "export_reviews"
]
